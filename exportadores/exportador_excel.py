# -*- coding: utf-8 -*-
# exportadores/exportador_excel.py

from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font


class ExportadorExcel:
    """
    Exporta resultados al libro final:
    - Atención de alertas con calidad (ALCON) -> A83
    - Histórico de Certificación de Gerentes  -> A180
    - TD Saldo (Temporales)                   -> A4
    - TD SABANA (conteo)                      -> E4
    - TD SABANA (DB)                          -> H4
    - TD SABANA (CR)                          -> K4
    """

    def __init__(self, ruta_salida: Path):
        self.ruta_salida = Path(ruta_salida)

    # 1) ALCON
    def exportar_atencion_alertas_calidad(
        self, df: pd.DataFrame, nombre_hoja_mes: str,
        nombre_archivo: str = "Indicadores_operacion_nuevo.xlsx",
        fila_inicio_excel: int = 83
    ) -> Path:
        destino = self.ruta_salida / nombre_archivo
        destino.parent.mkdir(parents=True, exist_ok=True)
        startrow = fila_inicio_excel - 1
        modo = "a" if destino.exists() else "w"
        with pd.ExcelWriter(destino, engine="openpyxl", mode=modo,
                            if_sheet_exists="replace" if modo == "a" else None) as writer:
            df.to_excel(writer, sheet_name=nombre_hoja_mes, index=False,
                        startrow=startrow, startcol=0)
        wb = load_workbook(destino); ws = wb[nombre_hoja_mes]
        first_data_row = startrow + 1
        data_start_row = first_data_row + 1
        data_end_row = data_start_row + len(df) - 1
        if len(df) > 0:
            col_index = 5
            for r in range(data_start_row, data_end_row + 1):
                cell = ws.cell(row=r, column=col_index)
                try:
                    if isinstance(cell.value, str) and cell.value.strip() != "":
                        cell.value = float(cell.value)
                except Exception:
                    pass
                cell.number_format = "0.00%"
        wb.save(destino)
        return destino

    # 2) HISTÓRICO
    def exportar_historico_certificacion(
        self, df: pd.DataFrame, nombre_hoja_mes: str,
        nombre_archivo: str = "Indicadores_operacion_nuevo.xlsx",
        fila_inicio_excel: int = 180, col_inicio_excel: int = 1,
        escribir_promedio: bool = True
    ) -> Path:
        destino = self.ruta_salida / nombre_archivo
        destino.parent.mkdir(parents=True, exist_ok=True)
        df = df.fillna("")
        wb = load_workbook(destino) if destino.exists() else Workbook()
        ws = wb[nombre_hoja_mes] if nombre_hoja_mes in wb.sheetnames else wb.create_sheet(title=nombre_hoja_mes)
        headers = list(df.columns)
        for j, h in enumerate(headers, start=col_inicio_excel):
            ws.cell(row=fila_inicio_excel, column=j, value=h)
        first_data_row = fila_inicio_excel + 1
        for i in range(len(df)):
            for j, h in enumerate(headers, start=col_inicio_excel):
                value = df.iloc[i][h]
                cell = ws.cell(row=first_data_row + i, column=j)
                if h == "INDICADOR" and value not in ("", None):
                    try:
                        num = float(value)
                        cell.value = num
                        cell.number_format = "0.00%"
                    except Exception:
                        cell.value = value
                else:
                    if pd.isna(value):
                        value = ""
                    cell.value = value
        if escribir_promedio:
            try:
                idx_ind = headers.index("INDICADOR")
            except ValueError:
                idx_ind = None
            if idx_ind is not None and len(df) > 0:
                def col_to_letter(col_num: int) -> str:
                    letters = ""; 
                    while col_num > 0:
                        col_num, remainder = divmod(col_num - 1, 26)
                        letters = chr(65 + remainder) + letters
                    return letters
                col_ind_excel = col_inicio_excel + idx_ind
                col_ind_letra = col_to_letter(col_ind_excel)
                last_data_row = first_data_row + len(df) - 1
                promedio_row = last_data_row + 1
                if col_ind_excel - 1 >= 1:
                    ws.cell(row=promedio_row, column=col_ind_excel - 1, value="Promedio INDICADOR")
                formula = f"=AVERAGE({col_ind_letra}{first_data_row}:{col_ind_letra}{last_data_row})"
                ws.cell(row=promedio_row, column=col_ind_excel, value=formula)
                ws.cell(row=promedio_row, column=col_ind_excel).number_format = "0.00%"
        wb.save(destino)
        return destino

    # Helper general
    def _col_letra(self, col_num: int) -> str:
        letters = ""
        while col_num > 0:
            col_num, rem = divmod(col_num - 1, 26)
            letters = chr(65 + rem) + letters
        return letters

    # 3) TD Saldo -> A4
    def exportar_td_saldo_en_indicadores(
        self, df_td: pd.DataFrame, nombre_hoja_mes: str,
        nombre_archivo: str = "Indicadores_operacion_nuevo.xlsx",
        celda_inicio: str = "A4"
    ) -> Path:
        destino = self.ruta_salida / nombre_archivo
        if not destino.exists():
            raise FileNotFoundError(f"No existe '{destino}'. Genéralo primero.")
        wb = load_workbook(destino)
        ws = wb[nombre_hoja_mes] if nombre_hoja_mes in wb.sheetnames else wb.create_sheet(title=nombre_hoja_mes)
        m = re.match(r"^([A-Za-z]+)(\d+)$", celda_inicio); 
        if not m: raise ValueError(f"Celda inicio inválida: {celda_inicio}")
        col_letters, row_str = m.groups()
        def col_letter_to_index(col: str) -> int:
            col = col.upper(); n = 0
            for ch in col: n = n*26 + (ord(ch) - 64)
            return n
        start_col = col_letter_to_index(col_letters); start_row = int(row_str)
        mapa = {
            "Etiquetas de fila": "Area",
            "Cuenta de SALDO CONTABLE": "TOTAL CUENTAS TEMPORALES CON SALDO",
            "Cuenta de VALOR PARTIDAS FUERA DE POLITICA": "CUENTAS TEMPORALES FUERA DE POLITICA",
        }
        df_es = df_td.rename(columns=mapa).copy()
        requeridas = ["Area", "TOTAL CUENTAS TEMPORALES CON SALDO", "CUENTAS TEMPORALES FUERA DE POLITICA"]
        for col in requeridas:
            if col not in df_es.columns:
                raise ValueError(f"Falta columna en TD Saldo: '{col}'")
        for r in range(start_row, start_row + 500):
            for c in range(start_col, start_col + 4):
                ws.cell(row=r, column=c, value=None)
        headers = ["Area", "TOTAL CUENTAS TEMPORALES CON SALDO", "CUENTAS TEMPORALES FUERA DE POLITICA", "%"]
        for j, h in enumerate(headers, start=start_col):
            ws.cell(row=start_row, column=j, value=h).font = Font(bold=True)
        first_data_row = start_row + 1
        for i in range(len(df_es)):
            area = df_es.iloc[i]["Area"]
            try: tot_cuentas = int(str(df_es.iloc[i]["TOTAL CUENTAS TEMPORALES CON SALDO"]).strip() or "0")
            except Exception: tot_cuentas = 0
            try: tot_fuera = int(str(df_es.iloc[i]["CUENTAS TEMPORALES FUERA DE POLITICA"]).strip() or "0")
            except Exception: tot_fuera = 0
            fila_excel = first_data_row + i
            ws.cell(row=fila_excel, column=start_col + 0, value=area)
            cB = ws.cell(row=fila_excel, column=start_col + 1, value=tot_cuentas); cB.number_format = "#,##0"
            cC = ws.cell(row=fila_excel, column=start_col + 2, value=tot_fuera);   cC.number_format = "#,##0"
            colB = self._col_letra(start_col + 1); colC = self._col_letra(start_col + 2)
            formula = f"=IFERROR({colC}{fila_excel}/{colB}{fila_excel},0)"
            ws.cell(row=fila_excel, column=start_col + 3, value=formula).number_format = "0.00%"
        wb.save(destino); return destino

    # 4) TD SABANA (conteo) -> E4
    def exportar_td_sabana_en_indicadores(
        self, df_td_sabana: pd.DataFrame, nombre_hoja_mes: str,
        nombre_archivo: str = "Indicadores_operacion_nuevo.xlsx",
        celda_inicio: str = "E4", formato_porcentaje: str = "0.0%"
    ) -> Path:
        destino = self.ruta_salida / nombre_archivo
        if not destino.exists():
            raise FileNotFoundError(f"No existe '{destino}'. Genéralo primero.")
        wb = load_workbook(destino)
        ws = wb[nombre_hoja_mes] if nombre_hoja_mes in wb.sheetnames else wb.create_sheet(title=nombre_hoja_mes)
        m = re.match(r"^([A-Za-z]+)(\d+)$", celda_inicio); 
        if not m: raise ValueError(f"Celda inicio inválida: {celda_inicio}")
        col_letters, row_str = m.groups()
        def col_letter_to_index(col: str) -> int:
            col = col.upper(); n = 0
            for ch in col: n = n*26 + (ord(ch) - 64)
            return n
        start_col = col_letter_to_index(col_letters); start_row = int(row_str)
        cols_map = {str(c).strip(): c for c in df_td_sabana.columns}
        col_total = cols_map.get("Total general")
        col_si = None
        for k in cols_map.keys():
            if str(k).strip().upper() in ("SI", "SÍ"):
                col_si = cols_map[k]; break
        if col_total is None or col_si is None:
            raise ValueError("TD SABANA (conteo) debe tener 'Total general' y 'SI'.")
        num_total = pd.to_numeric(df_td_sabana[col_total], errors="coerce")
        num_si    = pd.to_numeric(df_td_sabana[col_si],    errors="coerce")
        df_eff = df_td_sabana[(num_total.notna()) | (num_si.notna())].copy()
        max_filas_borrar = max(300, len(df_eff) + 30)
        for r in range(start_row, start_row + max_filas_borrar):
            for c in range(start_col, start_col + 3):
                ws.cell(row=r, column=c, value=None)
        headers = ["TOTAL PARTIDAS", "PARTIDAS FUERA DE POLITICA", "%"]
        for j, h in enumerate(headers, start=start_col):
            ws.cell(row=start_row, column=j, value=h).font = Font(bold=True)
        first_data_row = start_row + 1
        for i in range(len(df_eff)):
            fila_excel = first_data_row + i
            v_total = pd.to_numeric(df_eff.iloc[i][col_total], errors="coerce"); tot_part = int(float(v_total)) if pd.notna(v_total) else 0
            v_si = pd.to_numeric(df_eff.iloc[i][col_si], errors="coerce");         fuera    = int(float(v_si)) if pd.notna(v_si) else 0
            e_cell = ws.cell(row=fila_excel, column=start_col + 0, value=tot_part); e_cell.number_format = "#,##0"
            f_cell = ws.cell(row=fila_excel, column=start_col + 1, value=fuera);    f_cell.number_format = "#,##0"
            colE = self._col_letra(start_col + 0); colF = self._col_letra(start_col + 1)
            g_cell = ws.cell(row=fila_excel, column=start_col + 2, value=f"=IFERROR({colF}{fila_excel}/{colE}{fila_excel},0)")
            g_cell.number_format = formato_porcentaje
        wb.save(destino); return destino

    # 5) TD SABANA (DB) -> H4
    def exportar_td_sabana_db_en_indicadores(
        self, df_td_db: pd.DataFrame, nombre_hoja_mes: str,
        nombre_archivo: str = "Indicadores_operacion_nuevo.xlsx",
        celda_inicio: str = "H4", formato_porcentaje: str = "0.0%"
    ) -> Path:
        destino = self.ruta_salida / nombre_archivo
        if not destino.exists(): raise FileNotFoundError(f"No existe '{destino}'.")
        wb = load_workbook(destino); ws = wb[nombre_hoja_mes] if nombre_hoja_mes in wb.sheetnames else wb.create_sheet(title=nombre_hoja_mes)
        m = re.match(r"^([A-Za-z]+)(\d+)$", celda_inicio); 
        if not m: raise ValueError(f"Celda inicio inválida: {celda_inicio}")
        col_letters, row_str = m.groups()
        def col_letter_to_index(col: str) -> int:
            col = col.upper(); n = 0
            for ch in col: n = n*26 + (ord(ch) - 64)
            return n
        start_col = col_letter_to_index(col_letters); start_row = int(row_str)
        cols_map = {str(c).strip(): c for c in df_td_db.columns}
        col_total = cols_map.get("Total general"); col_si = None
        for k in cols_map.keys():
            if str(k).strip().upper() in ("SI", "SÍ"):
                col_si = cols_map[k]; break
        if col_total is None or col_si is None:
            raise ValueError("TD SABANA (DB) debe tener 'Total general' y 'SI'.")
        num_total = pd.to_numeric(df_td_db[col_total], errors="coerce")
        num_si    = pd.to_numeric(df_td_db[col_si],    errors="coerce")
        df_eff = df_td_db[(num_total.notna()) | (num_si.notna())].copy()
        max_filas_borrar = max(300, len(df_eff) + 30)
        for r in range(start_row, start_row + max_filas_borrar):
            for c in range(start_col, start_col + 3): ws.cell(row=r, column=c, value=None)
        headers = ["TOTAL VALOR PARTIDAS PESOS DB", "VALOR PARTIDAS PESOS DB (Fuera de política)", "%"]
        for j, h in enumerate(headers, start=start_col): ws.cell(row=start_row, column=j, value=h).font = Font(bold=True)
        first_data_row = start_row + 1
        for i in range(len(df_eff)):
            fila_excel = first_data_row + i
            v_total = pd.to_numeric(df_eff.iloc[i][col_total], errors="coerce"); tot_db = int(float(v_total)) if pd.notna(v_total) else 0
            v_si = pd.to_numeric(df_eff.iloc[i][col_si], errors="coerce");         si_db  = int(float(v_si)) if pd.notna(v_si) else 0
            h_cell = ws.cell(row=fila_excel, column=start_col + 0, value=tot_db); h_cell.number_format = "#,##0"
            i_cell = ws.cell(row=fila_excel, column=start_col + 1, value=si_db);  i_cell.number_format = "#,##0"
            colH = self._col_letra(start_col + 0); colI = self._col_letra(start_col + 1)
            j_cell = ws.cell(row=fila_excel, column=start_col + 2, value=f"=IFERROR({colI}{fila_excel}/{colH}{fila_excel},0)")
            j_cell.number_format = formato_porcentaje
        wb.save(destino); return destino

    # 6) TD SABANA (CR) -> K4
    def exportar_td_sabana_cr_en_indicadores(
        self, df_td_cr: pd.DataFrame, nombre_hoja_mes: str,
        nombre_archivo: str = "Indicadores_operacion_nuevo.xlsx",
        celda_inicio: str = "K4", formato_porcentaje: str = "0.0%"
    ) -> Path:
        destino = self.ruta_salida / nombre_archivo
        if not destino.exists(): raise FileNotFoundError(f"No existe '{destino}'.")
        wb = load_workbook(destino); ws = wb[nombre_hoja_mes] if nombre_hoja_mes in wb.sheetnames else wb.create_sheet(title=nombre_hoja_mes)
        m = re.match(r"^([A-Za-z]+)(\d+)$", celda_inicio); 
        if not m: raise ValueError(f"Celda inicio inválida: {celda_inicio}")
        col_letters, row_str = m.groups()
        def col_letter_to_index(col: str) -> int:
            col = col.upper(); n = 0
            for ch in col: n = n*26 + (ord(ch) - 64)
            return n
        start_col = col_letter_to_index(col_letters); start_row = int(row_str)
        cols_map = {str(c).strip(): c for c in df_td_cr.columns}
        col_total = cols_map.get("Total general"); col_si = None
        for k in cols_map.keys():
            if str(k).strip().upper() in ("SI", "SÍ"):
                col_si = cols_map[k]; break
        if col_total is None or col_si is None:
            raise ValueError("TD SABANA (CR) debe contener 'Total general' y 'SI'.")
        num_total = pd.to_numeric(df_td_cr[col_total], errors="coerce")
        num_si    = pd.to_numeric(df_td_cr[col_si],    errors="coerce")
        df_eff = df_td_cr[(num_total.notna()) | (num_si.notna())].copy()
        max_filas_borrar = max(300, len(df_eff) + 30)
        for r in range(start_row, start_row + max_filas_borrar):
            for c in range(start_col, start_col + 3): ws.cell(row=r, column=c, value=None)
        headers = ["TOTAL VALOR PARTIDAS PESOS CR", "VALOR PARTIDAS PESOS CR (Fuera de política)", "%"]
        for j, h in enumerate(headers, start=start_col): ws.cell(row=start_row, column=j, value=h).font = Font(bold=True)
        first_data_row = start_row + 1
        for i in range(len(df_eff)):
            fila_excel = first_data_row + i
            v_total = pd.to_numeric(df_eff.iloc[i][col_total], errors="coerce"); tot_cr = int(float(v_total)) if pd.notna(v_total) else 0
            v_si = pd.to_numeric(df_eff.iloc[i][col_si], errors="coerce");         si_cr  = int(float(v_si)) if pd.notna(v_si) else 0
            k_cell = ws.cell(row=fila_excel, column=start_col + 0, value=tot_cr); k_cell.number_format = "#,##0"
            l_cell = ws.cell(row=fila_excel, column=start_col + 1, value=si_cr);  l_cell.number_format = "#,##0"
            colK = self._col_letra(start_col + 0); colL = self._col_letra(start_col + 1)
            m_cell = ws.cell(row=fila_excel, column=start_col + 2, value=f"=IFERROR({colL}{fila_excel}/{colK}{fila_excel},0)")
            m_cell.number_format = formato_porcentaje
        wb.save(destino); return destino
