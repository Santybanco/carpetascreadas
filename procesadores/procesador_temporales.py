# -*- coding: utf-8 -*-
# procesadores/procesador_temporales.py

from pathlib import Path
from typing import Optional, Tuple, Dict
import re
import unicodedata
import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from datos.rutas import obtener_ruta_salida
from shutil import copy2


class ProcesadorTemporales:
    """
    Procesa el archivo 'Informe Cuentas Temporales ...xlsx':
      - Duplica 'TEMPORAL' -> 'TEMPORAL2' y filtra SALDO CONTABLE != 0
      - Genera 'TD Saldo' (conteos por gerencia)
      - Manipula 'Sábana Temporales' (DB/CR opcional) y construye TDs de SABANA:
            * Conteo (ya en hoja 'TD SABANA')
            * DB (bloque A32 dentro de 'TD SABANA')
            * CR (bloque A59 dentro de 'TD SABANA')
      - Añade fórmulas de control (SUM/SUBTOTAL) en 'Sábana Temporales'
      - Exporta a SharePoint (si existe ruta) o a datos/salida/ (guardado atómico)
    """

    def __init__(self, ruta_sharepoint: Optional[Path] = None):
        self.ruta_sharepoint = Path(ruta_sharepoint) if ruta_sharepoint else None

    # -------------------- UTILIDADES --------------------

    def _normalizar(self, s: str) -> str:
        if s is None:
            return ""
        s = re.sub(r"[\r\n\t]+", " ", str(s)).strip().lower()
        s = "".join(ch for ch in unicodedata.normalize("NFD", s)
                    if unicodedata.category(ch) != "Mn")
        return re.sub(r"\s+", " ", s)

    def _encontrar_columna(self, columnas: list, objetivo: str) -> Optional[str]:
        objetivo_n = self._normalizar(objetivo)
        mapa = {self._normalizar(c): c for c in columnas}
        return mapa.get(objetivo_n)

    def _to_numeric_safe(self, serie: pd.Series) -> pd.Series:
        if serie is None:
            return pd.Series(dtype=float)
        s = serie.astype(str)
        s = s.str.replace(r"[\r\n\t]", "", regex=True)
        s = s.str.replace(r"[^\d\-.,]", "", regex=True).str.strip()
        mask_coma = s.str.contains(",", regex=False)
        mask_punto = s.str.contains(r"\.", regex=True)
        s = s.where(~(mask_coma & mask_punto),
                    s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
        s = s.where(~(mask_coma & ~mask_punto),
                    s.str.replace(",", ".", regex=False))
        return pd.to_numeric(s, errors="coerce")

    def _limpiar_para_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        df2 = df.copy()
        for c in df2.select_dtypes(include=["object"]).columns:
            df2[c] = (
                df2[c].astype(str)
                .apply(lambda x: ILLEGAL_CHARACTERS_RE.sub("", x))
                .str.replace(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", regex=True)
            )
        return df2

    def _guardar_atomico(self, destino: Path, hojas: Dict[str, pd.DataFrame]) -> Path:
        destino.parent.mkdir(parents=True, exist_ok=True)
        tmp = destino.parent / f"{destino.stem}.__tmp__.xlsx"
        with pd.ExcelWriter(tmp, engine="openpyxl", mode="w") as writer:
            for nombre, df in hojas.items():
                self._limpiar_para_excel(df).to_excel(writer, sheet_name=nombre, index=False)
        if destino.exists():
            try:
                destino.unlink()
            except PermissionError:
                pass
        tmp.replace(destino)
        return destino

    def _duplicar_a_salida_local(self, origen: Path) -> Path:
        base_local = obtener_ruta_salida()
        base_local.mkdir(parents=True, exist_ok=True)
        destino_local = base_local / origen.name
        try:
            copy2(str(origen), str(destino_local))
        except Exception as e:
            print(f"⚠️ No se pudo copiar a salida local: {e}")
            return destino_local
        try:
            _ = pd.read_excel(destino_local, sheet_name=0, nrows=1, engine="openpyxl")
            print(
                f"🗂️ Copia local OK: {destino_local} "
                f"({round(destino_local.stat().st_size / 1024, 1)} KB)"
            )
        except Exception as e:
            print(f"⚠️ Copia local creada pero no legible: {e}")
        return destino_local

    # -------------------- TRANSFORMACIONES --------------------

    def _cargar_hoja(self, archivo: Path, nombre_hoja: str) -> pd.DataFrame:
        return pd.read_excel(archivo, sheet_name=nombre_hoja, dtype=str, engine="openpyxl")

    def _construir_temporal2(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, str, str, str]:
        col_saldo = self._encontrar_columna(df.columns, "SALDO CONTABLE")
        col_ger   = self._encontrar_columna(df.columns, "gerencia_responsable")
        col_fuera = self._encontrar_columna(df.columns, "PARTIDAS FUERA DE POLITICA_y")
        faltan = [n for n, v in {
            "SALDO CONTABLE": col_saldo,
            "gerencia_responsable": col_ger,
            "PARTIDAS FUERA DE POLITICA_y": col_fuera
        }.items() if v is None]
        if faltan:
            raise ValueError(f"No encontré columnas requeridas en 'TEMPORAL': {', '.join(faltan)}")
        saldo = self._to_numeric_safe(df[col_saldo]).fillna(0)
        df2 = df[saldo != 0].copy()
        df2[col_ger] = df2[col_ger].astype(str).str.replace(r"[\r\n\t]", "", regex=True).str.strip()
        return df2, col_ger, col_saldo, col_fuera

    def _construir_td_saldo(self, df2: pd.DataFrame, col_ger: str, col_saldo: str, col_fuera: str) -> pd.DataFrame:
        td1 = df2.groupby(col_ger, dropna=False).size().reset_index(name="Cuenta de SALDO CONTABLE")
        flag_fuera = self._to_numeric_safe(df2[col_fuera]).fillna(0).ne(0)
        td2 = flag_fuera.groupby(df2[col_ger]).sum().reset_index(name="Cuenta de VALOR PARTIDAS FUERA DE POLITICA")
        out = td1.merge(td2, on=col_ger, how="left").fillna(0)
        out.rename(columns={col_ger: "Etiquetas de fila"}, inplace=True)
        out.loc[len(out)] = [
            "Total general",
            int(out["Cuenta de SALDO CONTABLE"].sum()),
            int(out["Cuenta de VALOR PARTIDAS FUERA DE POLITICA"].sum())
        ]
        return out

    def _agregar_db_cr_sabana(self, df_sabana: pd.DataFrame) -> pd.DataFrame:
        cols = list(df_sabana.columns)
        col_valor = self._encontrar_columna(cols, "VALOR PARTIDA PESOS")
        if not col_valor:
            raise ValueError("No encontré 'VALOR PARTIDA PESOS' en 'Sábana Temporales'.")
        serie = self._to_numeric_safe(df_sabana[col_valor]).fillna(0.0)
        db = serie.where(serie > 0, 0.0)
        cr = serie.where(serie < 0, 0.0)
        df_out = df_sabana.copy()
        for c in ["VALOR PARTIDA PESOS DB", "VALOR PARTIDA PESOS CR"]:
            if c in df_out.columns:
                df_out = df_out.drop(columns=[c])
        pos = cols.index(col_valor) + 1
        df_out.insert(pos,     "VALOR PARTIDA PESOS DB", db)
        df_out.insert(pos + 1, "VALOR PARTIDA PESOS CR", cr)
        suma_L  = float(serie.sum())
        suma_DB = float(db.sum())
        suma_CR = float(cr.sum())
        delta   = (suma_DB + suma_CR) - suma_L
        if abs(delta) > 0.01:
            print("⚠️ Aviso Sábana: L vs M+N difiere.",
                  f"L={suma_L:,.2f} | M={suma_DB:,.2f} | N={suma_CR:,.2f} | Δ={delta:,.2f}")
        else:
            print("✅ Sumas OK Sábana:",
                  f"L={suma_L:,.2f} == M+N={suma_DB + suma_CR:,.2f} (Δ={delta:,.2f})")
        return df_out

    def _construir_td_sabana(self, df: pd.DataFrame) -> pd.DataFrame:
        col_g = self._encontrar_columna(df.columns, "gerencia_responsable")
        col_p = self._encontrar_columna(df.columns, "FUERA DE POLITICA")
        col_v = self._encontrar_columna(df.columns, "VALOR PARTIDA PESOS")
        faltan = [n for n, v in {
            "gerencia_responsable": col_g, "FUERA DE POLITICA": col_p, "VALOR PARTIDA PESOS": col_v
        }.items() if v is None]
        if faltan:
            raise ValueError(f"No encontré columnas para TD SABANA: {', '.join(faltan)}")
        dfp = df[[col_g, col_p, col_v]].copy()
        dfp[col_g] = dfp[col_g].astype(str).str.replace(r"[\r\n\t]", "", regex=True).str.strip()
        dfp[col_p] = (dfp[col_p].astype(str)
                      .str.replace(r"[\r\n\t]", "", regex=True).str.strip().str.upper()
                      .replace({"SÍ": "SI"}))
        td = pd.pivot_table(dfp, index=col_g, columns=col_p, values=col_v,
                            aggfunc="count", fill_value=0, dropna=False)
        orden_cols = [c for c in ["NO", "SI"] if c in td.columns]
        td = td[orden_cols] if orden_cols else td
        td["Total general"] = td.sum(axis=1)
        td.loc["Total general"] = td.sum()
        return td.reset_index().rename(columns={col_g: "Etiquetas de fila"})

    # ---------- NUEVOS (DB ya lo tenías; añadimos CR + verificación) ----------

    def _construir_td_sabana_db(self, df_sabana: pd.DataFrame) -> pd.DataFrame:
        cols = list(df_sabana.columns)
        col_g = self._encontrar_columna(cols, "gerencia_responsable")
        col_p = self._encontrar_columna(cols, "FUERA DE POLITICA")
        col_db = self._encontrar_columna(cols, "VALOR PARTIDA PESOS DB")
        col_v  = self._encontrar_columna(cols, "VALOR PARTIDA PESOS")
        faltan = [n for n, v in {"gerencia_responsable": col_g, "FUERA DE POLITICA": col_p}.items() if v is None]
        if faltan:
            raise ValueError(f"No encontré columnas requeridas para TD SABANA DB: {', '.join(faltan)}")
        dfp = df_sabana.copy()
        if col_db is None:
            if col_v is None:
                raise ValueError("No encontré 'VALOR PARTIDA PESOS DB' ni 'VALOR PARTIDA PESOS'.")
            serie = self._to_numeric_safe(dfp[col_v]).fillna(0.0)
            dfp["__DB__"] = serie.where(serie > 0, 0.0)
            col_db = "__DB__"
        dfp[col_g] = dfp[col_g].astype(str).str.replace(r"[\r\n\t]", "", regex=True).str.strip()
        dfp[col_p] = (dfp[col_p].astype(str)
                      .str.replace(r"[\r\n\t]", "", regex=True).str.strip().str.upper()
                      .replace({"SÍ": "SI"}))
        td = pd.pivot_table(dfp, index=col_g, columns=col_p, values=col_db,
                            aggfunc="sum", fill_value=0.0, dropna=False)
        cols_ord = [c for c in ["SI", "NO"] if c in td.columns]
        td = td[cols_ord] if cols_ord else td
        td["Total general"] = td.sum(axis=1)
        totales = {c: float(td[c].sum()) for c in td.columns}
        td.loc["Total general"] = totales
        return td.reset_index().rename(columns={col_g: "Etiquetas de fila"})

    def _construir_td_sabana_cr(self, df_sabana: pd.DataFrame) -> pd.DataFrame:
        cols = list(df_sabana.columns)
        col_ger = self._encontrar_columna(cols, "gerencia_responsable")
        col_pol = self._encontrar_columna(cols, "FUERA DE POLITICA")
        col_cr  = self._encontrar_columna(cols, "VALOR PARTIDA PESOS CR")
        faltan = [n for n, v in {
            "gerencia_responsable": col_ger, "FUERA DE POLITICA": col_pol, "VALOR PARTIDA PESOS CR": col_cr
        }.items() if v is None]
        if faltan:
            raise ValueError(f"No encontré columnas requeridas para TD SABANA CR: {', '.join(faltan)}")
        dfp = df_sabana[[col_ger, col_pol, col_cr]].copy()
        dfp[col_ger] = dfp[col_ger].astype(str).str.replace(r"[\r\n\t]", "", regex=True).str.strip()
        dfp[col_pol] = (dfp[col_pol].astype(str)
                        .str.replace(r"[\r\n\t]", "", regex=True).str.strip().str.upper()
                        .replace({"SÍ": "SI"}))
        dfp[col_cr] = self._to_numeric_safe(dfp[col_cr]).fillna(0.0)
        td = pd.pivot_table(dfp, index=col_ger, columns=col_pol, values=col_cr,
                            aggfunc="sum", fill_value=0.0, dropna=False)
        cols_ord = [c for c in ["SI", "NO"] if c in td.columns]
        td = td[cols_ord] if cols_ord else td
        td["Total general"] = td.sum(axis=1)
        totales = {c: float(td[c].sum()) for c in td.columns}
        td.loc["Total general"] = totales
        return td.reset_index().rename(columns={col_ger: "Etiquetas de fila"})

    def _escribir_bloque_td_sabana_db(self, xlsx_path: Path, df_db: pd.DataFrame, celda_inicio: str = "A32") -> None:
        xlsx_path = Path(xlsx_path)
        wb = load_workbook(xlsx_path)
        hoja = "TD SABANA"
        if hoja not in wb.sheetnames:
            wb.save(xlsx_path); return
        ws = wb[hoja]
        m = re.match(r"^([A-Za-z]+)(\d+)$", celda_inicio)
        if not m:
            wb.save(xlsx_path); raise ValueError(f"Celda inicio inválida: {celda_inicio}")
        col_letters, row_str = m.groups()
        def col_letter_to_index(col: str) -> int:
            col = col.upper(); n = 0
            for ch in col: n = n*26 + (ord(ch)-64)
            return n
        start_col = col_letter_to_index(col_letters)
        start_row = int(row_str)
        cols_map = {str(c).strip(): c for c in df_db.columns}
        has_si = "SI" in cols_map
        headers = ["Etiquetas de fila"] + (["SI"] if has_si else []) + ["Total general"]
        max_rows = max(500, len(df_db) + 20)
        for r in range(start_row, start_row + max_rows):
            for c in range(start_col, start_col + 3):
                ws.cell(row=r, column=c, value=None)
        for j, h in enumerate(headers, start=start_col):
            ws.cell(row=start_row, column=j, value=h)
        first_data_row = start_row + 1
        for i in range(len(df_db)):
            fila_excel = first_data_row + i
            ws.cell(row=fila_excel, column=start_col, value=str(df_db.iloc[i]["Etiquetas de fila"]))
            col_w = start_col + 1
            if has_si:
                val_si = float(pd.to_numeric(df_db.iloc[i]["SI"], errors="coerce") or 0.0)
                cSI = ws.cell(row=fila_excel, column=col_w, value=val_si)
                cSI.number_format = "#,##0.00"; col_w += 1
            val_tot = float(pd.to_numeric(df_db.iloc[i]["Total general"], errors="coerce") or 0.0)
            cTOT = ws.cell(row=fila_excel, column=col_w, value=val_tot)
            cTOT.number_format = "#,##0.00"
        wb.save(xlsx_path)

    def _escribir_bloque_td_sabana_cr(self, xlsx_path: Path, df_cr: pd.DataFrame, celda_inicio: str = "A59") -> None:
        xlsx_path = Path(xlsx_path)
        wb = load_workbook(xlsx_path)
        hoja = "TD SABANA"
        if hoja not in wb.sheetnames:
            wb.save(xlsx_path); return
        ws = wb[hoja]
        m = re.match(r"^([A-Za-z]+)(\d+)$", celda_inicio)
        if not m:
            wb.save(xlsx_path); raise ValueError(f"Celda inicio inválida: {celda_inicio}")
        col_letters, row_str = m.groups()
        def col_letter_to_index(col: str) -> int:
            col = col.upper(); n = 0
            for ch in col: n = n*26 + (ord(ch)-64)
            return n
        start_col = col_letter_to_index(col_letters)
        start_row = int(row_str)
        cols_map = {str(c).strip(): c for c in df_cr.columns}
        has_si = "SI" in cols_map or "SÍ" in cols_map
        col_si = cols_map.get("SI", cols_map.get("SÍ"))
        headers = ["Etiquetas de fila"] + (["SI"] if has_si else []) + ["Total general"]
        max_rows = max(500, len(df_cr) + 20)
        for r in range(start_row, start_row + max_rows):
            for c in range(start_col, start_col + 3):
                ws.cell(row=r, column=c, value=None)
        for j, h in enumerate(headers, start=start_col):
            ws.cell(row=start_row, column=j, value=h)
        first_data_row = start_row + 1
        for i in range(len(df_cr)):
            fila_excel = first_data_row + i
            ws.cell(row=fila_excel, column=start_col, value=str(df_cr.iloc[i]["Etiquetas de fila"]))
            col_w = start_col + 1
            if has_si:
                val_si = float(pd.to_numeric(df_cr.iloc[i][col_si], errors="coerce") or 0.0)
                cSI = ws.cell(row=fila_excel, column=col_w, value=val_si)
                cSI.number_format = "#,##0.00"; col_w += 1
            val_tot = float(pd.to_numeric(df_cr.iloc[i]["Total general"], errors="coerce") or 0.0)
            cTOT = ws.cell(row=fila_excel, column=col_w, value=val_tot)
            cTOT.number_format = "#,##0.00"
        wb.save(xlsx_path)

    def _verificar_cr_por_gerencia(self, df_sabana: pd.DataFrame, td_cr: pd.DataFrame, tolerancia: float = 0.01) -> None:
        cols = list(df_sabana.columns)
        col_ger = self._encontrar_columna(cols, "gerencia_responsable")
        col_pol = self._encontrar_columna(cols, "FUERA DE POLITICA")
        col_cr  = self._encontrar_columna(cols, "VALOR PARTIDA PESOS CR")
        if not all([col_ger, col_pol, col_cr]):
            print("⚠️ Verificación CR omitida: columnas no disponibles."); return
        dfp = df_sabana[[col_ger, col_pol, col_cr]].copy()
        dfp[col_ger] = dfp[col_ger].astype(str).str.replace(r"[\r\n\t]", "", regex=True).str.strip()
        dfp[col_pol] = (dfp[col_pol].astype(str)
                        .str.replace(r"[\r\n\t]", "", regex=True).str.strip().str.upper()
                        .replace({"SÍ": "SI"}))
        cr = self._to_numeric_safe(dfp[col_cr]).fillna(0.0)
        dfp["_CR_"] = cr
        suma_si = dfp[dfp[col_pol].eq("SI")].groupby(col_ger, dropna=False)["_CR_"].sum()
        td = td_cr.copy()
        td = td[td["Etiquetas de fila"].astype(str).str.strip().str.lower() != "total general"]
        col_si_td = "SI" if "SI" in td.columns else ("SÍ" if "SÍ" in td.columns else None)
        if not col_si_td:
            print("⚠️ Verificación CR: la TD no trae columna SI/SÍ; omito chequeo."); return
        problemas = []
        for _, row in td.iterrows():
            ger = str(row["Etiquetas de fila"]).strip()
            td_val = float(pd.to_numeric(row[col_si_td], errors="coerce") or 0.0)
            base_val = float(suma_si.get(ger, 0.0))
            if abs(td_val - base_val) > tolerancia:
                problemas.append((ger, td_val, base_val, td_val - base_val))
        if problemas:
            print("❌ Verificación CR por gerencia: se encontraron diferencias:")
            for ger, td_val, base_val, delta in problemas[:30]:
                print(f"   - {ger}: TD(SI)={td_val:,.2f} vs Base(SI)={base_val:,.2f}  Δ={delta:,.2f}")
            if len(problemas) > 30:
                print(f"   ... y {len(problemas)-30} gerencias más.")
        else:
            print(f"✅ Verificación CR por gerencia: {len(suma_si)} gerencias sin diferencias (> {tolerancia}).")

    def _agregar_filas_control_excel_sabana(self, xlsx_path: Path, hoja: str = "Sábana Temporales") -> None:
        xlsx_path = Path(xlsx_path)
        wb = load_workbook(xlsx_path)
        if hoja not in wb.sheetnames:
            wb.save(xlsx_path); return
        ws = wb[hoja]
        headers = {str(ws.cell(row=1, column=c).value).strip(): c for c in range(1, ws.max_column + 1)}
        cL = headers.get("VALOR PARTIDA PESOS")
        cM = headers.get("VALOR PARTIDA PESOS DB")
        cN = headers.get("VALOR PARTIDA PESOS CR")
        if not all([cL, cM, cN]):
            wb.save(xlsx_path); return
        L = get_column_letter(cL); M = get_column_letter(cM); N = get_column_letter(cN)
        first_data = 2; last_data  = ws.max_row; total_row  = last_data + 1; diff_row   = last_data + 2
        ws.cell(row=total_row, column=1, value="TOTAL SUMA")
        ws.cell(row=diff_row,  column=1, value="M+N - L (debe ser 0)")
        func = 9
        ws.cell(row=total_row, column=cL, value=f"=SUM({L}{first_data}:{L}{last_data})")
        ws.cell(row=total_row, column=cM, value=f"=SUBTOTAL({func},{M}{first_data}:{M}{last_data})")
        ws.cell(row=total_row, column=cN, value=f"=SUBTOTAL({func},{N}{first_data}:{N}{last_data})")
        ws.cell(row=diff_row, column=cL,
                value=f"=SUBTOTAL({func},{M}{first_data}:{M}{last_data})+SUBTOTAL({func},{N}{first_data}:{N}{last_data})-SUM({L}{first_data}:{L}{last_data})")
        for r in (total_row, diff_row):
            for c in (cL, cM, cN):
                ws.cell(row=r, column=c).number_format = "#,##0.00"
        wb.save(xlsx_path)

    # -------------------- ORQUESTADOR --------------------

    def procesar_y_exportar(self, archivo_temporales: Path, crear_db_cr_sabana: bool = True) -> Path:
        archivo_temporales = Path(archivo_temporales)

        print("1) Cargando hojas TEMPORAL y Sábana Temporales…")
        df_temporal = self._cargar_hoja(archivo_temporales, "TEMPORAL")
        df_sabana   = self._cargar_hoja(archivo_temporales, "Sábana Temporales")

        print("2) Construyendo TEMPORAL2 (SALDO CONTABLE != 0)…")
        df_temporal2, col_ger, col_saldo, col_fuera = self._construir_temporal2(df_temporal)

        print("3) Construyendo TD Saldo…")
        td_saldo = self._construir_td_saldo(df_temporal2, col_ger, col_saldo, col_fuera)

        if crear_db_cr_sabana:
            print("4) Agregando DB/CR (M y N) en Sábana…")
            df_sabana_proc = self._agregar_db_cr_sabana(df_sabana)
        else:
            print("4) (Omitido) DB/CR en Sábana por configuración.")
            df_sabana_proc = df_sabana

        print("5) Construyendo TD SABANA DB (SUMA de VALOR PARTIDA PESOS DB)…")
        td_sabana_db = self._construir_td_sabana_db(df_sabana_proc)

        print("6) Construyendo TD SABANA (conteo)…")
        td_sabana = self._construir_td_sabana(df_sabana)

        print("7) Construyendo TD SABANA CR (SUMA de VALOR PARTIDA PESOS CR)…")
        td_sabana_cr = self._construir_td_sabana_cr(df_sabana_proc)

        base_dest = self.ruta_sharepoint if (self.ruta_sharepoint and self.ruta_sharepoint.exists()) else obtener_ruta_salida()
        if base_dest == obtener_ruta_salida():
            print(f"⚠️ SharePoint no disponible. Guardando en salida local: {base_dest}")

        destino = base_dest / f"{archivo_temporales.stem}_procesado.xlsx"

        destino = self._guardar_atomico(destino, {
            "TEMPORAL": df_temporal,
            "TEMPORAL2": df_temporal2,
            "TD Saldo": td_saldo,
            "Sábana Temporales": df_sabana_proc,
            "TD SABANA": td_sabana
        })

        # Posprocesado en el XLSX
        self._agregar_filas_control_excel_sabana(destino, hoja="Sábana Temporales")
        self._escribir_bloque_td_sabana_db(destino, td_sabana_db, celda_inicio="A32")
        self._escribir_bloque_td_sabana_cr(destino, td_sabana_cr, celda_inicio="A59")
        self._verificar_cr_por_gerencia(df_sabana_proc, td_sabana_cr, tolerancia=0.01)

        # Backup local desactivado durante pruebas (actívalo cuando quieras)
        # self._duplicar_a_salida_local(destino)

        try:
            _ = pd.read_excel(destino, sheet_name="TEMPORAL", nrows=1, engine="openpyxl")
            print(f"✅ Verificación OK: {destino.name} se puede leer.")
        except Exception as e:
            print(f"❌ Alerta: el archivo se guardó pero no se pudo leer con pandas: {e}")

        return destino
